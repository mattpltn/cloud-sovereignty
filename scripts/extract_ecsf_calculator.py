#!/usr/bin/env python3
"""Phase 2b.1: extracts data/extracted/ecsf-calculator.json (the 48
"specific objectives" of the official Sovereignty assessment calculator
XLSX, one sheet, "Sov Assessment Levels (COL)"), plus
data/local/ecsf-calculator-verbatim.json (git-ignored, D-009/D-010
regime).

Parses the sheet algorithmically (not hand-transcribed) so blank/
malformed rows are detected mechanically rather than guessed: the sheet
uses merged cells in the original workbook that openpyxl reports as
None on every row but the merge's top-left, which for several answer
rows leaves a "phantom" row carrying only a SEAL value, or a label-only/
value-only row. Per instruction, these are captured as-is (missing
fields simply omitted) with needs_review, not repaired or inferred.

The sheet's own "Score" column is a worked example the sheet's own
header note (column I, row 4) states is "filled with fictitious values
... for the sole purpose of exemplification" — it is never extracted.

Run: .venv/bin/python3 scripts/extract_ecsf_calculator.py
"""

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "sources" / "Annex - Sovereignty assessment calculator.xlsx"
OUT_PUBLIC = ROOT / "data" / "extracted" / "ecsf-calculator.json"
OUT_LOCAL = ROOT / "data" / "local" / "ecsf-calculator-verbatim.json"

DOCUMENT = "Annex - Sovereignty assessment calculator.xlsx (EU Commission, Cloud III DPS tender annex)"
SHEET = "Sov Assessment Levels (COL)"
FR_PENDING = "FR-TRANSLATION-PENDING"
LOCAL_VERBATIM_PLACEHOLDER = "SEE-LOCAL-VERBATIM"

DOMAIN_RE = re.compile(r"^(SOV-\d)\b")

# Provisional, unverified coverage hints: each calculator specific
# objective -> the v1.2.1 contributing factor(s) (ecsf.json ids) it
# operationalizes. Built by reading the calculator's objective titles
# against the guidance PDF's own "Criteria" bullets (section "Assessment
# of Sovereignty Objectives", pages 3-5), which are near-verbatim the
# same bullets extracted as ecsf.json factors in Phase 2b. Several
# objectives split a single v1.2.1 bullet into two calculator rows (e.g.
# SOV-1 objectives 2/3 both operationalize csat-sov1-ecsf-02); SOV-7 and
# SOV-8 have no ecsf.json records at all (inheritance-only / out of
# scope, Phase 2b), so every SOV-7/SOV-8 objective is "uncovered" by
# construction, not by absence of a plausible match.
FACTOR_HINTS = {
    "SOV-1": {1: ["csat-sov1-ecsf-01"], 2: ["csat-sov1-ecsf-02"], 3: ["csat-sov1-ecsf-02"],
              4: ["csat-sov1-ecsf-03"], 5: ["csat-sov1-ecsf-04"], 6: ["csat-sov1-ecsf-05"],
              7: ["csat-sov1-ecsf-05"], 8: ["csat-sov1-ecsf-06"]},
    "SOV-2": {1: ["csat-sov2-ecsf-01"], 2: ["csat-sov2-ecsf-02"], 3: ["csat-sov2-ecsf-03"],
              4: ["csat-sov2-ecsf-04"], 5: ["csat-sov2-ecsf-05"], 6: ["csat-sov2-ecsf-05"]},
    "SOV-3": {1: ["csat-sov3-ecsf-01"], 2: ["csat-sov3-ecsf-02"], 3: ["csat-sov3-ecsf-02"],
              4: ["csat-sov3-ecsf-03"], 5: ["csat-sov3-ecsf-04"]},
    "SOV-4": {1: ["csat-sov4-ecsf-01"], 2: ["csat-sov4-ecsf-02"], 3: ["csat-sov4-ecsf-03"],
              4: ["csat-sov4-ecsf-04"], 5: ["csat-sov4-ecsf-05"], 6: ["csat-sov4-ecsf-06"]},
    "SOV-5": {1: ["csat-sov5-ecsf-01"], 2: ["csat-sov5-ecsf-01"], 3: ["csat-sov5-ecsf-02"],
              4: ["csat-sov5-ecsf-03"], 5: ["csat-sov5-ecsf-03"], 6: ["csat-sov5-ecsf-04"],
              7: ["csat-sov5-ecsf-05"]},
    "SOV-6": {1: ["csat-sov6-ecsf-01"], 2: ["csat-sov6-ecsf-01"], 3: ["csat-sov6-ecsf-02"],
              4: ["csat-sov6-ecsf-03"], 5: ["csat-sov6-ecsf-04"]},
    "SOV-7": {i: [] for i in range(1, 8)},
    "SOV-8": {i: [] for i in range(1, 5)},
}
UNCOVERED_NOTE_NO_RECORDS = (
    "SOV-7/SOV-8 have no ecsf.json control records (inheritance-only / "
    "out of scope per CLAUDE.md and Phase 2b) — coverage: uncovered by "
    "construction, not a judgment that no analogue exists."
)


def ls(en: str) -> dict:
    return {"en": en, "fr": FR_PENDING}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]

    objectives = []
    current_domain = None
    current_domain_weight = None
    current_obj = None

    def flush_obj():
        if current_obj is not None:
            objectives.append(current_obj)

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        a, b, c, d, e, f = (cell.value for cell in row[:6])

        if isinstance(a, str) and DOMAIN_RE.match(a):
            flush_obj()
            current_obj = None
            current_domain = DOMAIN_RE.match(a).group(1)
            current_domain_weight = d
            continue

        if isinstance(a, int):
            flush_obj()
            current_obj = {
                "sov_domain": current_domain,
                "objective_number": a,
                "description_en": (b or "").strip(),
                "row_start": row[0].row,
                "row_end": row[0].row,
                "answers": [],
            }
            # fall through: this row's C/D/F may also carry answer 1

        if current_obj is None:
            continue  # stray row before the first domain header

        current_obj["row_end"] = row[0].row
        if c is None and d is None and f is None:
            continue  # pure blank spacer row, no data at all

        answer = {}
        review_notes = []
        if c is None:
            review_notes.append("blank answer label in source row")
        elif isinstance(c, str) and c.strip() == "":
            answer["label_en"] = c
            review_notes.append("whitespace-only answer label in source row")
        elif not isinstance(c, str):
            answer["label_en"] = str(c)
            review_notes.append(f"non-text answer label in source row (raw value: {c!r})")
        else:
            answer["label_en"] = c

        if d is None:
            review_notes.append("missing point value in source row")
        else:
            answer["point_value"] = d

        if f is None:
            review_notes.append("missing SEAL value in source row")
        else:
            answer["seal_level"] = f

        if review_notes:
            answer["needs_review"] = True
            answer["needs_review_note"] = "; ".join(review_notes)
        current_obj["answers"].append(answer)

    flush_obj()

    public_records = []
    local_verbatim = {}

    for obj in objectives:
        dom = obj["sov_domain"]
        num = obj["objective_number"]
        oid = f"{dom.lower().replace('-', '')}-so{num}"

        desc_key = f"{oid}-desc"
        local_verbatim[desc_key] = ls(obj["description_en"])

        answer_options = []
        obj_needs_review = False
        for i, ans in enumerate(obj["answers"], start=1):
            opt = {}
            if "label_en" in ans:
                label_key = f"{oid}-opt{i}"
                local_verbatim[label_key] = ls(ans["label_en"])
                opt["label"] = {"en": LOCAL_VERBATIM_PLACEHOLDER, "fr": FR_PENDING}
            if "point_value" in ans:
                opt["point_value"] = ans["point_value"]
            if "seal_level" in ans:
                opt["seal_level"] = ans["seal_level"]
            if ans.get("needs_review"):
                opt["needs_review"] = True
                opt["needs_review_note"] = ans["needs_review_note"]
                obj_needs_review = True
            answer_options.append(opt)

        hints_list = FACTOR_HINTS.get(dom, {}).get(num, [])
        coverage = "covered" if hints_list else "uncovered"
        hints_entry = {"coverage": coverage, "ecsf_factor_candidates": hints_list}

        record = {
            "id": oid,
            "sov_domain": dom,
            "objective_number": num,
            "description": {"en": LOCAL_VERBATIM_PLACEHOLDER, "fr": FR_PENDING},
            "answer_options": answer_options,
            "ecsf_factor_hints": hints_entry,
            "source_pointer": {
                "document": DOCUMENT,
                "section": f"{SHEET}, rows {obj['row_start']}-{obj['row_end']}",
            },
        }
        if obj_needs_review:
            record["needs_review"] = True
        if dom in ("SOV-7", "SOV-8"):
            record.setdefault("needs_review", True)
            record["needs_review_note"] = UNCOVERED_NOTE_NO_RECORDS
        public_records.append(record)

    OUT_PUBLIC.parent.mkdir(parents=True, exist_ok=True)
    OUT_LOCAL.parent.mkdir(parents=True, exist_ok=True)
    OUT_PUBLIC.write_text(json.dumps(public_records, indent=2, ensure_ascii=False) + "\n")
    OUT_LOCAL.write_text(json.dumps(local_verbatim, indent=2, ensure_ascii=False) + "\n")
    print(f"Wrote {len(public_records)} specific objectives -> {OUT_PUBLIC}")
    print(f"Wrote {len(local_verbatim)} verbatim entries -> {OUT_LOCAL}")


if __name__ == "__main__":
    main()
